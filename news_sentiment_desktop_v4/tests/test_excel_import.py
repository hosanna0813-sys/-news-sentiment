"""測試：Excel 多工作表讀取（規格五：不可只讀第一張工作表）"""
from __future__ import annotations

import openpyxl
from app.services.importer.excel_importer import import_file


def _make_multi_sheet_excel(path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "政治"
    ws1.append(["文章標題", "摘要", "來源", "時間", "作者", "網址"])
    ws1.append(["政治新聞A", "摘要A", "中央社", "2026-01-01", "記者甲", "http://example.com/a"])

    ws2 = wb.create_sheet("財經")
    ws2.append(["title", "summary", "source", "published_at", "url"])
    ws2.append(["財經新聞B", "摘要B", "經濟日報", "2026-01-02", "http://example.com/b"])

    ws3 = wb.create_sheet("社會")
    ws3.append(["標題", "來源", "時間"])
    ws3.append(["社會新聞C", "自由時報", "2026-01-03"])

    wb.save(str(path))


def test_multi_sheet_all_read(tmp_path):
    excel_path = tmp_path / "multi_sheet.xlsx"
    _make_multi_sheet_excel(excel_path)

    result = import_file(str(excel_path))

    assert result.sheet_count == 3
    assert result.total_rows == 3
    titles = {it.title for it in result.items}
    assert titles == {"政治新聞A", "財經新聞B", "社會新聞C"}
    # 每個工作表都要被記錄在來源分頁
    sheets_seen = {it.source_sheet for it in result.items}
    assert sheets_seen == {"政治", "財經", "社會"}
