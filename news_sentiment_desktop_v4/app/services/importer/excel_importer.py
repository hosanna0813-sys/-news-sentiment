"""
新聞匯入服務 — 對應規格書 五、匯入新聞

支援 Excel (.xlsx，所有工作表) 與 CSV。
自動辨識並標準化欄位（含 KEYPO 常見欄位別名）。
每一列都建立全新的唯一 row_id（uuid4），即使標題/網址/內容完全重複也不衝突。
"""
from __future__ import annotations

import csv
import time
from pathlib import Path
from typing import List, Dict, Any, Optional

from app.models.news import NewsItem
from app.utils.text_utils import new_id, normalize_whitespace
from app.utils.logging_setup import get_logger

logger = get_logger("importer")

# 欄位別名對應表（含 KEYPO 常見欄位）
FIELD_ALIASES: Dict[str, List[str]] = {
    "news_id": ["news_id", "id", "編號", "新聞編號"],
    "title": ["title", "標題", "文章標題", "新聞標題"],
    "summary": ["summary", "摘要", "文章摘要"],
    "body": ["body", "正文", "內文", "全文", "content", "文章內容"],
    "source": ["source", "來源", "媒體", "媒體來源", "media"],
    "published_at": ["published_at", "時間", "發布時間", "刊登時間", "日期", "publish_date", "date"],
    "author": ["author", "作者", "記者"],
    "url": ["url", "網址", "連結", "link", "原始連結"],
    "channel": ["channel", "頻道", "版面"],
    "tags": ["tags", "標籤", "分類"],
}


def _build_reverse_alias_map() -> Dict[str, str]:
    rev = {}
    for std_field, aliases in FIELD_ALIASES.items():
        for a in aliases:
            rev[a.strip().lower()] = std_field
    return rev


_REVERSE_ALIAS = _build_reverse_alias_map()


def map_columns(raw_columns: List[str]) -> Dict[str, str]:
    """回傳 {原始欄名: 標準欄名}，辨識不到的欄位保留原名（存入 tags 或忽略）"""
    mapping = {}
    for col in raw_columns:
        key = str(col).strip().lower()
        std = _REVERSE_ALIAS.get(key)
        if std:
            mapping[col] = std
    return mapping


class ImportResult:
    def __init__(self):
        self.import_batch_id = new_id("batch_")
        self.file_name = ""
        self.sheet_count = 0
        self.total_rows = 0
        self.duplicate_rows = 0
        self.missing_url_rows = 0
        self.has_body_rows = 0
        self.summary_only_rows = 0
        self.column_mapping_summary: Dict[str, Dict[str, str]] = {}
        self.items: List[NewsItem] = []


def import_file(file_path: str) -> ImportResult:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"找不到檔案: {file_path}")
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return _import_excel(path)
    elif ext == ".csv":
        return _import_csv(path)
    else:
        raise ValueError(f"不支援的檔案格式: {ext}（僅支援 .xlsx / .csv）")


def _import_excel(path: Path) -> ImportResult:
    import openpyxl  # 延遲 import，避免未安裝時影響其他功能

    result = ImportResult()
    result.file_name = path.name
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    result.sheet_count = len(wb.sheetnames)

    for sheet_name in wb.sheetnames:  # 規格要求：必須讀取所有工作表，不可只讀第一張
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue
        header = [str(h).strip() if h is not None else "" for h in header]
        col_map = map_columns(header)
        result.column_mapping_summary[sheet_name] = col_map

        for raw_row in rows_iter:
            if raw_row is None or all(v is None for v in raw_row):
                continue
            row_dict = {header[i]: raw_row[i] for i in range(len(header)) if i < len(raw_row)}
            item = _row_to_news_item(row_dict, col_map, result.import_batch_id, sheet_name)
            if item is None:
                continue
            result.items.append(item)
    wb.close()
    _finalize_result(result)
    return result


def _import_csv(path: Path) -> ImportResult:
    result = ImportResult()
    result.file_name = path.name
    result.sheet_count = 1
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        _finalize_result(result)
        return result
    header = [str(h).strip() for h in rows[0]]
    col_map = map_columns(header)
    result.column_mapping_summary["CSV"] = col_map
    for raw_row in rows[1:]:
        if not raw_row or all(not v for v in raw_row):
            continue
        row_dict = {header[i]: raw_row[i] for i in range(len(header)) if i < len(raw_row)}
        item = _row_to_news_item(row_dict, col_map, result.import_batch_id, "CSV")
        if item is None:
            continue
        result.items.append(item)
    _finalize_result(result)
    return result


def _row_to_news_item(row_dict: Dict[str, Any], col_map: Dict[str, str],
                       import_batch_id: str, sheet_name: str) -> Optional[NewsItem]:
    std: Dict[str, str] = {}
    for orig_col, std_field in col_map.items():
        val = row_dict.get(orig_col)
        std[std_field] = normalize_whitespace(str(val)) if val is not None else ""

    title = std.get("title", "")
    if not title:
        # 沒有標題的列視為無效資料，略過但不中斷整體匯入
        return None

    body = std.get("body", "")
    published_at = std.get("published_at", "")
    if hasattr(published_at, "isoformat"):
        published_at = published_at.isoformat()  # type: ignore
    else:
        published_at = str(published_at)

    return NewsItem(
        row_id=new_id("row_"),
        news_id=std.get("news_id") or "",
        import_batch_id=import_batch_id,
        source_sheet=sheet_name,
        title=title,
        summary=std.get("summary", ""),
        source=std.get("source", ""),
        published_at=published_at,
        author=std.get("author", ""),
        url=std.get("url", ""),
        channel=std.get("channel", ""),
        tags=std.get("tags", ""),
        excel_body=body,
        body_text=body,
        body_source="Excel正文" if body else "無正文",
        body_word_count=len(body),
        retained=True,
        retention_status="待確認",
    )


def _finalize_result(result: ImportResult) -> None:
    result.total_rows = len(result.items)
    result.missing_url_rows = sum(1 for it in result.items if not it.url)
    result.has_body_rows = sum(1 for it in result.items if it.excel_body)
    result.summary_only_rows = sum(1 for it in result.items if it.summary and not it.excel_body)

    # O(n) 分組偵測重複：以 (標題小寫去空白, 網址小寫去空白) 為 key
    groups: Dict[str, List[NewsItem]] = {}
    for it in result.items:
        key = f"{it.title.strip().lower()}|{it.url.strip().lower()}"
        groups.setdefault(key, []).append(it)

    dup_count = 0
    for key, members in groups.items():
        if len(members) > 1:
            group_id = new_id("dup_")
            for it in members:
                it.duplicate_group_id = group_id
                dup_count += 1
        else:
            members[0].duplicate_group_id = ""
    result.duplicate_rows = dup_count
    logger.info(f"匯入完成: {result.file_name}, 共 {result.total_rows} 筆, "
                f"{result.sheet_count} 個工作表, 重複 {result.duplicate_rows} 筆")
